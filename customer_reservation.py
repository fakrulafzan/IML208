from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os 
from tkinter.ttk import Combobox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Customer Reservation System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('customer_record.xlsx')
if not file.exists():
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Reservation No."
    sheet['B1'] = "Cust Name"
    sheet['C1'] = "Cust IC Num"
    sheet['D1'] = "Cust DOB"
    sheet['E1'] = "Cust Phone No."
    sheet['F1'] = "Cust Email"
    sheet['G1'] = "Cust Address"
    sheet['H1'] = "Check In"
    sheet['I1'] = "Check Out"
    sheet['J1'] = "Pax"
    sheet['K1'] = "Homestay Type"
    sheet['L1'] = "Payment Type"
    sheet['M1'] = "Price"

    workbook.save('customer_record.xlsx')


# Exit window
def Exit():
    root.destroy()

#Resrvation No.
def reservation_no():
    file= openpyxl.load_workbook('customer_record.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row, column=1).value
    
    try:
        Reservation.set(max_row_value+1)
    except:
        Reservation.set("20231")

#clear
def Clear():
    Name.set('')
    IC.set('')
    DOB.set('')
    Phone.set('')
    Email.set('')
    Address.set('')
    C_in.set('')
    C_out.set('')
    Pax.set('')
    Homestay_Type.set("Select Homestay")
    Payment_Type.set("Select Payment")
    Price.set('')

    

# save function
def Save():
    try:
        # Get data from input fields
        reservation_id = Reservation.get()
        name = Name.get()
        ic_number = IC.get()
        dob = DOB.get()
        phone_number = Phone.get()
        email_address = Email.get()
        address = Address.get()
        check_in_date = C_in.get()
        check_out_date = C_out.get()
        pax = Pax.get()
        homestay_type = Homestay_Type.get()
        payment_type = Payment_Type.get()
        price = Price.get()

        # Check for missing data
        missing_data = False
        for field in [reservation_id, name, ic_number, dob, phone_number, email_address, address, check_in_date, check_out_date, pax, homestay_type, payment_type, price]:
            if field == "" or field == "Select Pax" or field == "Select Homestay" or field == "Select Payment":
                missing_data = True
                break

        if missing_data:
            messagebox.showerror("Error", "Missing data! Please fill in all required fields.")
        else:
            # Open workbook
            workbook = openpyxl.load_workbook('customer_record.xlsx')
            sheet = workbook.active

            # Write data to new row
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1).value = reservation_id
            sheet.cell(row=row, column=2).value = name
            sheet.cell(row=row, column=3).value = ic_number
            sheet.cell(row=row, column=4).value = dob
            sheet.cell(row=row, column=5).value = phone_number
            sheet.cell(row=row, column=6).value = email_address
            sheet.cell(row=row, column=7).value = address
            sheet.cell(row=row, column=8).value = check_in_date
            sheet.cell(row=row, column=9).value = check_out_date
            sheet.cell(row=row, column=10).value = pax
            sheet.cell(row=row, column=11).value = homestay_type
            sheet.cell(row=row, column=12).value = payment_type
            sheet.cell(row=row, column=13).value = price

            # Save workbook
            workbook.save(r'customer_record.xlsx')

            # Show success message
            messagebox.showinfo("Success", "Data saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Unexpected error occurred: {e}")

##search button
def search():
    
    text = Search.get()

    Clear()
    saveButton.config(state='disable')

    file=openpyxl.load_workbook("customer_record.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
##          print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

##            print(reg_no_position)
##            print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid reservation number!!!")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    x13=sheet.cell(row=int(reg_number),column=13).value

    #print(x1)
   # print(x3)
   # print(x4)
   # print(x5)
   # print(x6)
   # print(x7)
   # print(x8)
   # print(x9)
    #print(x10)
    #print(x11)
   # print(x12)
   # print(x13)

    Reservation.set(x1)
    Name.set(x2)
    IC.set(x3)
    DOB.set(x4)
    Phone.set(x5)
    Email.set(x6)
    Address.set(x7)
    C_in.set(x8)
    C_out.set(x9)
    Pax.set(x10)
    Homestay_Type.set(x11)
    Payment_Type.set(x12)
    Price.set(x13)

def Update(filename, reservation_id, new_data):
    """
    Updates existing reservation information in a spreadsheet.

    Args:
        filename: Path to the spreadsheet file.
        reservation_id: ID of the reservation to update.
        new_data: Dictionary containing updated data for the reservation.
    """
    try:
        # Open workbook
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Get row index for reservation ID
        row_index = None
        for row in sheet.rows:
            if row[0].value == reservation_id:
                row_index = row[0].row
                break

        # Check if reservation found
        if not row_index:
            raise ValueError(f"Reservation ID '{reservation_id}' not found.")

        # Update data
        for column, value in new_data.items():
            sheet.cell(row=row_index, column=column + 1).value = value

        # Save workbook
        workbook.save(filename)

        # Show success message
        messagebox.showinfo("Success", "Data updated successfully!")
    except Exception as e:
        # Log error
        print(f"Error updating reservation '{reservation_id}': {e}")
        # Show error message
        messagebox.showerror("Error", f"Unexpected error occurred: {e}")


# top frames
Label(root, text="Email: fakrulafzan19@gmail.com", width=10, height=3, bg="lightgrey", anchor='e').pack(side=TOP, fill=X)
Label(root, text="CUSTOMER RESERVATION", width=10, height=2, bg="beige", fg='black', font='arial 20 bold').pack(side=TOP, fill=X)

# search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)

imageicon3 = PhotoImage(file="Images/search_icon.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='lightgrey', font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="Images/refresh.png")
Update_button = Button(root, image=imageicon4, bg="lightgrey")
Update_button.place(x=110, y=66)

#Reservation and Date
Label(root,text="Reservation No:",font="arial 13", fg=framebg, bg=background).place (x=30,y=150)
Label(root,text="Date",font="arial 13", fg=framebg, bg=background).place (x=500,y=150)

Reservation=IntVar()
Date = StringVar()

resr_entry = Entry(root,textvariable=Reservation, width=15,font="arial 18")
resr_entry.place(x=160, y=150)


reservation_no()

today = date.today()
dl = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable= Date, width=15, font="arial 10")
date_entry.place(x=550,y=150)

Date.set(dl)

#Customer details
obj=LabelFrame(root,text="Customer's Details", font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=50)
Label(obj,text="IC Num.:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=100)
Label(obj,text="Date of Birth:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=150)
Label(obj,text="Phone No.:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=50)
Label(obj,text="Email:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=100)
Label(obj,text="Address:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=150)

Name=StringVar()
name_entry = Entry(obj,textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160,y=50)

IC = StringVar()
ic_entry = Entry(obj,textvariable=IC, width=20, font="arial 10")
ic_entry.place(x=160,y=100)

DOB = StringVar()
dob_entry = Entry(obj,textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160,y=150)

Phone = StringVar()
phone_entry = Entry(obj,textvariable=Phone, width=20, font="arial 10")
phone_entry.place(x=630,y=50)

Email = StringVar()
email_entry = Entry(obj,textvariable=Email, width=20, font="arial 10")
email_entry.place(x=630,y=100)

Address = StringVar()
address_entry = Entry(obj,textvariable=Address, width=20, font="arial 10")
address_entry.place(x=630,y=150)

#Homestay details
obj2=LabelFrame(root,text="Booking Details", font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Check In:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=50)
Label(obj2,text="Check Out:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=100)
Label(obj2,text="Pax:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=150)
Label(obj2,text="Homestay Type:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=50)
Label(obj2,text="Payment Type:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=100)
Label(obj2,text="Price:", font="arial 13", bg=framebg,fg=framefg).place(x=500, y=150)

C_in = StringVar()
checkin_entry = Entry(obj2,textvariable=C_in, width=20, font="arial 10")
checkin_entry.place(x=200,y=50)

C_out = StringVar()
checkout_entry = Entry(obj2,textvariable=C_out, width=20, font="arial 10")
checkout_entry.place(x=200,y=100)

Pax = StringVar()
px_entry = Entry(obj2,textvariable=Pax, width=20, font="arial 10")
px_entry.place(x=200,y=150)

Homestay_Type = Combobox(obj2, values=['LA STANDARD', 'LA PRIVALLEY', 'LA PRIVALLEY (PREMIUM)','LA PRIMA VOUGE','LA PRINCE', 'LA PRINCE VICE', 'LA SYA','LA MUNI', 'LA VIDAS','LA SYAMUNI PREMIUM CLASS'],font="Roboto 10", width=17, state="r")
Homestay_Type.place(x=630, y=50)
Homestay_Type.set("Select Homestay")

Payment_Type = Combobox(obj2, values=['Cash', 'QR payment', 'Online Transfer','Credit/Debit','E-Wallet'],font="Roboto 10", width=17, state="r")
Payment_Type.place(x=630, y=100)
Payment_Type.set("Select Payment")

Price = StringVar()
price_entry = Entry(obj2,textvariable=Price, width=20, font="arial 10")
price_entry.place(x=630,y=150)

#image
f=Frame(root,bd=3,bg="beige", width=200,height=280,relief=GROOVE)
f.place(x=1000,y=200)

img=PhotoImage(file="Images/pricelist.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button
saveButton=Button(root, text="Save", width=19, height=1, font="arial 12 bold", bg="lightgrey", command=Save)
saveButton.place(x=1000, y=500)
Button(root, text="Reset", width=19, height=1, font="arial 12 bold", bg="lightgrey", command=Clear).place(x=1000, y=550)
Button(root, text="Exit", width=19, height=1, font="arial 12 bold", bg="lightgrey", command=Exit).place(x=1000, y=600)


root.mainloop()
